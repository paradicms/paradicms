import logging
from io import BytesIO
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook


# def _convert_csv_row_to_json_object(csv_row_dict: Dict[str, Any]) -> Dict[str, Any]:
#     """
#     Convert a CSV row {header: cell} dict to a JSON object.
#
#     Headers separators (".") indicate nested values.
#     Arrays are not supported.
#     """
#
#     root_json_object = {}
#     for csv_key, csv_value in csv_row_dict.items():
#         if csv_value is None:
#             continue
#         key_split = csv_key.split(".")
#         if len(key_split) == 1:
#             root_json_object[csv_key] = csv_value
#             continue
#         json_object = root_json_object
#         for sub_key in key_split[:-1]:
#             json_object = json_object.setdefault(sub_key, {})
#         json_object[key_split[-1]] = csv_value
#     return root_json_object


class Excel2010Extractor:
    def __init__(self, *, xlsx_file_path: Path):
        self.__logger = logging.getLogger(__name__)
        self.__xlsx_file_path = xlsx_file_path

    def __call__(self, **kwds):
        openpyxl_workbook = load_workbook(str(self.__xlsx_file_path), data_only=True)
        try:
            openpyxl_workbook.iso_dates = True
            sheets = {}
            for openpyxl_sheet_name in openpyxl_workbook.sheetnames:
                openpyxl_sheet = openpyxl_workbook[openpyxl_sheet_name]

                images_by_row_col = {}
                for openpyxl_image in openpyxl_sheet._images:
                    image = Image.open(BytesIO(openpyxl_image._data())).copy()
                    images_by_row_col.setdefault(openpyxl_image.anchor._from.row, {})[
                        openpyxl_image.anchor._from.col
                    ] = image

                rows = []
                for row_i, openpyxl_row in enumerate(openpyxl_sheet.rows):
                    row = []
                    for column_i, openpyxl_cell in enumerate(openpyxl_row):
                        try:
                            row.append(images_by_row_col[row_i][column_i])
                        except KeyError:
                            row.append(openpyxl_cell.value)
                    rows.append(tuple(row))
                sheets[openpyxl_sheet_name] = {"rows": tuple(rows)}
            return {"sheets": sheets}
        finally:
            openpyxl_workbook.close()
