from pathlib import Path

from more_itertools import consume
from openpyxl import load_workbook

from paradicms_etl.extractors.excel_2010_extractor import Excel2010Extractor
from paradicms_etl.loaders.excel_2010_loader import Excel2010Loader
from paradicms_etl.transformers.spreadsheet_transformer import SpreadsheetTransformer


def test_load(excel_2010_test_data_file_path: Path, tmp_path: Path):
    models = tuple(
        SpreadsheetTransformer(pipeline_id="test")(
            **Excel2010Extractor(xlsx_file_path=excel_2010_test_data_file_path)(
                force=False
            )
        )
    )
    out_excel_2010_file_path = tmp_path / "test.xlsx"
    loader = Excel2010Loader(xlsx_file_path=out_excel_2010_file_path)
    consume(loader(flush=True, models=models))
    assert out_excel_2010_file_path.is_file()
    workbook = load_workbook(str(out_excel_2010_file_path), data_only=True)
    assert "FoafPerson" in workbook.sheetnames
